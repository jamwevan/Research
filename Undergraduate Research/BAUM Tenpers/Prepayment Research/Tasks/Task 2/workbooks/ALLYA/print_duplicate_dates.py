import sys
import re
from datetime import datetime, date
from typing import Optional, Dict, List

from openpyxl import load_workbook


def normalize_date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v

    s = str(v).strip()
    if s == "":
        return None

    s = s.replace(",", "").strip()

    fmts = [
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%Y/%m/%d",
        "%d-%b-%Y",
        "%d-%b-%y",
        "%b %d %Y",
        "%b %d %y",
    ]
    for fmt in fmts:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass

    digits = re.sub(r"\D", "", s)
    if len(digits) == 8:
        for fmt in ("%Y%m%d", "%m%d%Y"):
            try:
                return datetime.strptime(digits, fmt).date()
            except ValueError:
                pass

    return None


def main():
    if len(sys.argv) < 2:
        print("Usage: python print_duplicate_dates.py <workbook.xlsx>")
        sys.exit(1)

    path = sys.argv[1]
    wb = load_workbook(path, data_only=True)

    any_dups = False

    for ws in wb.worksheets:
        seen: Dict[date, int] = {}
        dups: Dict[date, List[int]] = {}

        # Assume header in row 1; dates start at row 2
        for r in range(2, ws.max_row + 1):
            d = normalize_date(ws.cell(row=r, column=1).value)  # Column A
            if d is None:
                continue

            if d in seen:
                dups.setdefault(d, [seen[d]]).append(r)
            else:
                seen[d] = r

        if dups:
            any_dups = True
            print(f"\nSheet: {ws.title}")
            for d in sorted(dups.keys()):
                rows = dups[d]
                print(f"  {d.isoformat()}  (rows {rows})")

    if not any_dups:
        print("No duplicate dates found in column A.")


if __name__ == "__main__":
    main()
