from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def is_empty_cell_value(v) -> bool:
    """Treat None, empty string, or whitespace-only string as empty."""
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False

def dedupe_sheet(ws, header_rows=1):
    """
    Removes duplicate data rows in a worksheet.
    - Keeps the first `header_rows` rows untouched.
    - Deduping key is the full row values across all columns (1..max_column).
    """
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row <= header_rows:
        return 0

    seen = set()
    rows_to_delete = []

    for r in range(header_rows + 1, max_row + 1):
        row_values = tuple(ws.cell(row=r, column=c).value for c in range(1, max_col + 1))
        if row_values in seen:
            rows_to_delete.append(r)
        else:
            seen.add(row_values)

    for r in reversed(rows_to_delete):
        ws.delete_rows(r, 1)

    return len(rows_to_delete)

def highlight_empty_cells_A_to_M(ws, header_rows=1):
    """
    Highlights empty cells in columns A..M (1..13) in yellow.
    Applies to all rows (including headers by default); set header_rows to skip headers.
    """
    max_row = ws.max_row
    start_row = 1 if header_rows == 0 else 1  # keep simple: highlight everywhere unless you want to skip
    # If you want to SKIP headers, change to: start_row = header_rows + 1
    for r in range(start_row, max_row + 1):
        for c in range(1, 14):  # A(1) .. M(13)
            cell = ws.cell(row=r, column=c)
            if is_empty_cell_value(cell.value):
                cell.fill = YELLOW_FILL

def process_workbook(input_path, output_path=None, header_rows=1):
    input_path = Path(input_path)
    if output_path is None:
        output_path = input_path.with_name(input_path.stem + "_deduped_highlighted" + input_path.suffix)

    wb = load_workbook(input_path)

    total_deleted = 0
    per_sheet_deleted = {}

    for ws in wb.worksheets:
        deleted = dedupe_sheet(ws, header_rows=header_rows)
        per_sheet_deleted[ws.title] = deleted
        total_deleted += deleted

        highlight_empty_cells_A_to_M(ws, header_rows=header_rows)

    wb.save(output_path)
    return output_path, total_deleted, per_sheet_deleted

if __name__ == "__main__":
    in_file = "SDART.xlsx"  # <-- change if needed

    out_file, total_deleted, per_sheet = process_workbook(
        in_file,
        header_rows=1  # assumes row 1 is a header row
    )

    print(f"Saved: {out_file}")
    print(f"Total duplicate rows deleted: {total_deleted}")
    print("Per sheet (deleted duplicates):")
    for name, cnt in per_sheet.items():
        print(f"  {name}: {cnt}")
